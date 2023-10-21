import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("lab5.xlsx")
ws = wb.active

blueFill = PatternFill(fill_type='solid', fgColor='00fff7')
greenFill = PatternFill(fill_type='solid', fgColor='008000')

ws.conditional_formatting.add('H4:Z7',
                              CellIsRule(operator='equal', formula=['H$4'], stopIfTrue=True, fill=blueFill))
ws.conditional_formatting.add('H4:Z7',
                              CellIsRule(operator='equal', formula=['I$5'], stopIfTrue=True, fill=greenFill))
wb.save('копия.xlsx')
